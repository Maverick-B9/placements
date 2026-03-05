
import openpyxl, json, re

TIMES = [
    "10:00 AM \u2013 11:00 AM",
    "11:00 AM \u2013 12:00 PM",
    "12:00 PM \u2013 1:00 PM",
    "2:00 PM \u2013 3:00 PM",
    "3:00 PM \u2013 4:00 PM"
]

def parse_dob(dob):
    if not dob: return ''
    s = str(dob).strip()
    parts = re.split(r'[\/\-\.\s]+', s)
    nums = [p for p in parts if p.isdigit()]
    if len(nums) == 3:
        d, m, y = nums
        if len(y) == 2: y = '20' + y
        if len(y) == 4:
            return f"{d.zfill(2)}/{m.zfill(2)}/{y}"
    return s

# Parse Excel
wb = openpyxl.load_workbook('Placement_Drive_Schedule (2).xlsx', data_only=True)
ws = wb['Master Student Schedule']

students, companies_set = [], []

# Skip the first two header rows based on inspection
for row in ws.iter_rows(min_row=3, values_only=True):
    cols = (list(row) + [None]*16)[:16]
    si, name, usn, branch, gender, dob, p10, p12, cgpa, resume, c1, c2, c3, c4, c5, status_ok = cols
    if not usn: continue
    dob_str = parse_dob(dob)
    password = dob_str if dob_str else '01/01/2000'
    schedule = []
    # Sessions are in columns 11-15 (indexed 10-14)
    for i, comp in enumerate([c1, c2, c3, c4, c5]):
        cname = str(comp).strip() if comp else ''
        schedule.append({"round": i+1, "company": cname, "time": TIMES[i], "status": "pending", "remark": "", "result": ""})
        if cname and cname not in companies_set:
            companies_set.append(cname)
    students.append({
        "id": int(si) if si else len(students)+1,
        "usn": str(usn).strip(), "name": str(name).strip() if name else '',
        "branch": str(branch).strip() if branch else '',
        "email": '', # Email not found in Master Student Schedule sheet but was in Session sheets
        "phone": '', # Phone not found in Master Student Schedule sheet but was in Session sheets
        "gender": str(gender).strip() if gender else '',
        "dob": dob_str, "pct10": str(p10) if p10 else '',
        "pct12": str(p12) if p12 else '', "cgpa": str(cgpa) if cgpa else '',
        "resume": str(resume).strip() if resume else '',
        "schedule": schedule, "password": password
    })

companies_list = sorted(companies_set)
print(f"Parsed: {len(students)} students, {len(companies_list)} companies")

# Parse Room Allotment Excel for company -> room number mapping
raw_room_map = {}
try:
    rwb = openpyxl.load_workbook('Room Allotment_Master.xlsx', data_only=True)
    for sn in rwb.sheetnames:
        rws = rwb[sn]
        for row in rws.iter_rows(min_row=2, max_row=rws.max_row, values_only=True):
            cols = list(row)
            company = cols[1] if len(cols) > 1 else None
            room = cols[3] if len(cols) > 3 else None
            if company and room:
                raw_room_map[str(company).strip()] = str(room).strip()
    print(f"Parsed: {len(raw_room_map)} room allotments")
except Exception as e:
    print(f"Warning: Could not parse Room Allotment file: {e}")

# Match room allotment company names to schedule company names (fuzzy matching)
def normalize(s):
    return re.sub(r'[^a-z0-9]', '', s.lower())

room_map = {}
# Build a lookup from normalized room-allotment names
norm_room = {normalize(k): (k, v) for k, v in raw_room_map.items()}
matched = 0
for sched_name in companies_list:
    # Try exact match first
    if sched_name in raw_room_map:
        room_map[sched_name] = raw_room_map[sched_name]
        matched += 1
    else:
        # Try normalized match
        nk = normalize(sched_name)
        if nk in norm_room:
            room_map[sched_name] = norm_room[nk][1]
            matched += 1
print(f"Matched {matched}/{len(companies_list)} schedule companies to room numbers")

# Build JSON constants (safe — no JS braces here)
raw_json  = json.dumps(students, ensure_ascii=False, separators=(',', ':'))
comp_json = json.dumps(companies_list, ensure_ascii=False, separators=(',', ':'))
room_json = json.dumps(room_map, ensure_ascii=False, separators=(',', ':'))
data_block = f"const RAW={raw_json};\nconst COMPANY={comp_json};\nconst ROOM_DATA_INIT={room_json};\n"

# Read JS logic from separate files (avoids f-string/JS brace collision)
with open('cloud_store.js', 'r', encoding='utf-8') as f:
    cloud_store = f.read()

with open('portal_logic.js', 'r', encoding='utf-8') as f:
    js_logic = f.read()

with open('company_logic.js', 'r', encoding='utf-8') as f:
    company_logic = f.read()


full_script = data_block + "\n" + cloud_store + "\n" + js_logic + "\n" + company_logic

# Read and update HTML
with open('index.html', 'r', encoding='utf-8') as f:
    html = f.read()

# Update credentials table in login page
old_start = html.find('<table class="creds-table">')
old_end   = html.find('</table>', old_start) + len('</table>')
if old_start != -1:
    rows = "\n".join(
        f"<tr><td class='us'>{c.lower().replace(' ','')[:20]}</td><td class='mu'>{c}</td></tr>"
        for c in companies_list
    )
    new_table = (
        "<table class=\"creds-table\">\n"
        "<tr><td><strong>Login ID</strong></td><td><strong>Company Name</strong></td></tr>\n"
        + rows + "\n</table>"
    )
    html = html[:old_start] + new_table + html[old_end:]

# Replace entire <script>…</script> block using regex for robustness
pattern = re.compile(r'<script>.*?</script>', re.DOTALL)
new_script_tag = f"<script>\n{full_script}\n</script>"

# Add Firebase SDK scripts before the closing </head> if not already present
firebase_sdk = '''    <script src="https://www.gstatic.com/firebasejs/10.12.0/firebase-app-compat.js"></script>
    <script src="https://www.gstatic.com/firebasejs/10.12.0/firebase-firestore-compat.js"></script>'''
if 'firebase-app-compat' not in html:
    html = html.replace('</head>', firebase_sdk + '\n</head>')

# Add loading overlay after <body> if not already present
loading_overlay = '''    <div id="loading-overlay" style="position:fixed;top:0;left:0;width:100%;height:100%;background:var(--bg,#090d1a);display:flex;align-items:center;justify-content:center;z-index:9999;flex-direction:column;gap:12px">
        <div style="width:36px;height:36px;border:3px solid rgba(99,102,241,.2);border-top:3px solid #6366f1;border-radius:50%;animation:spin 1s linear infinite"></div>
        <div style="font-family:'DM Sans',sans-serif;font-size:12px;color:#5c6494;letter-spacing:1px">LOADING PORTAL DATA...</div>
        <style>@keyframes spin{to{transform:rotate(360deg)}}</style>
    </div>'''
if 'id="loading-overlay"' not in html:
    html = html.replace('<body>', '<body>\n' + loading_overlay)
if pattern.search(html):
    # Use a lambda for the replacement to avoid backslash issues (\s, \1, etc.)
    html = pattern.sub(lambda _: new_script_tag, html)
else:
    # Fallback if tag not found as expected (should not happen if file is valid)
    print("Warning: <script> tag not found with standard pattern, attempting fallback...")
    ss = html.find('<script>')
    se = html.rfind('</script>')
    if ss != -1 and se != -1:
        html = html[:ss] + new_script_tag + html[se + len('</script>'):]

with open('index.html', 'w', encoding='utf-8') as f:
    f.write(html)

print("Portal rebuilt successfully!")
print(f"File size: {len(html.encode('utf-8')) / 1024:.0f} KB")
