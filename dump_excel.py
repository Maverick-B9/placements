
import openpyxl, json
wb = openpyxl.load_workbook('Master Copy of Students for Placement Drive.xlsx', data_only=True)
ws = wb.active
headers = [cell.value for cell in ws[1]]
first_row = [cell.value for cell in ws[2]]
print("HEADERS:")
for i, h in enumerate(headers):
    print(f"{i}: {h}")
print("\nFIRST ROW:")
for i, v in enumerate(first_row):
    print(f"{i}: {v}")
