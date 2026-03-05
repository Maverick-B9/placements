
import openpyxl, sys

wb = openpyxl.load_workbook('Placement_Drive_Schedule (2).xlsx', data_only=True)

with open('excel_info.txt', 'w', encoding='utf-8') as out:
    out.write(f"Sheets: {wb.sheetnames}\n\n")
    for sname in wb.sheetnames:
        ws = wb[sname]
        out.write(f"=== {sname} === rows={ws.max_row} cols={ws.max_column}\n")
        for r in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            out.write(str(r) + '\n')
        out.write('\n')

print("Done. Check excel_info.txt")
