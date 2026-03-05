
import openpyxl, json, re

TIMES = [
    "10:00 AM \u2013 11:00 AM",
    "11:00 AM \u2013 12:00 PM",
    "12:00 PM \u2013 1:00 PM",
    "2:00 PM \u2013 3:00 PM",
    "3:00 PM \u2013 4:00 PM"
]

# Parse Excel (Day 2)
wb = openpyxl.load_workbook('MITM_Placement_Fair_2026_Schedule (1).xlsx', data_only=True)
ws = wb['Master Student Schedule']

# ── Cross-reference Day 1 Excel for resume hyperlinks ──
# Day 1 Excel has resume links as hyperlink objects in column [9]
day1_resume_map = {}  # USN (upper) → Google Drive URL
try:
    wb1 = openpyxl.load_workbook('Placement_Drive_Schedule (2).xlsx')  # NOT data_only, need hyperlinks
    ws1 = wb1['Master Student Schedule']
    for row in ws1.iter_rows(min_row=3):
        cells = list(row)
        usn_val = cells[2].value if len(cells) > 2 else None
        if not usn_val: continue
        resume_cell = cells[9] if len(cells) > 9 else None
        if resume_cell and resume_cell.hyperlink and resume_cell.hyperlink.target:
            day1_resume_map[str(usn_val).strip().upper()] = resume_cell.hyperlink.target
        elif resume_cell and resume_cell.value and str(resume_cell.value).strip().startswith('http'):
            day1_resume_map[str(usn_val).strip().upper()] = str(resume_cell.value).strip()
    print(f"Day 1 resume cross-ref: {len(day1_resume_map)} resume links loaded")
except Exception as e:
    print(f"Warning: Could not load Day 1 resume data: {e}")

students, companies_set = [], []

# Day 2 Excel has 13 columns (Row 2 = headers, data from Row 3):
# [0] SI NO, [1] Name, [2] USN, [3] Branch, [4] Email, [5] Phone,
# [6] CGPA, [7] S1:10-11AM, [8] S2:11-12PM, [9] S3:12-1PM,
# [10] S4:2-3PM, [11] S5:3-4PM, [12] Status
for row in ws.iter_rows(min_row=3, values_only=True):
    cols = (list(row) + [None]*13)[:13]
    si, name, usn, branch, email, phone, cgpa, c1, c2, c3, c4, c5, status_ok = cols
    if not usn: continue
    # No DOB in Day 2 Excel — use default password
    password = '01/01/2004'

    # Look up resume from Day 1 by USN
    usn_key = str(usn).strip().upper()
    resume_url = day1_resume_map.get(usn_key, '')

    schedule = []
    for i, comp in enumerate([c1, c2, c3, c4, c5]):
        cname = str(comp).strip() if comp else ''
        schedule.append({"round": i+1, "company": cname, "time": TIMES[i], "status": "pending", "remark": "", "result": ""})
        if cname and cname not in companies_set:
            companies_set.append(cname)
    students.append({
        "id": int(si) if si else len(students)+1,
        "usn": str(usn).strip(), "name": str(name).strip() if name else '',
        "branch": str(branch).strip() if branch else '',
        "email": str(email).strip() if email else '',
        "phone": str(phone).strip() if phone else '',
        "gender": '',
        "dob": '', "pct10": '',
        "pct12": '', "cgpa": str(cgpa) if cgpa else '',
        "resume": resume_url,
        "schedule": schedule, "password": password
    })

companies_list = sorted(companies_set)
print(f"Day 2 Parsed: {len(students)} students, {len(companies_list)} companies")

# Parse Room Allotment Excel for company -> room number mapping
raw_room_map = {}
try:
    rwb = openpyxl.load_workbook('Room Allotment_Master.xlsx', data_only=True)
    for sn in rwb.sheetnames:
        rws = rwb[sn]
        for row in rws.iter_rows(min_row=2, max_row=rws.max_row, values_only=True):
            cols = list(row)
            company = cols[1] if len(cols) > 1 else None
            room = cols[3] if len(cols) > 3 else None
            if company and room:
                raw_room_map[str(company).strip()] = str(room).strip()
    print(f"Parsed: {len(raw_room_map)} room allotments")
except Exception as e:
    print(f"Warning: Could not parse Room Allotment file: {e}")

# Match room allotment company names to schedule company names (fuzzy matching)
def normalize(s):
    return re.sub(r'[^a-z0-9]', '', s.lower())

room_map = {}
norm_room = {normalize(k): (k, v) for k, v in raw_room_map.items()}
matched = 0
for sched_name in companies_list:
    if sched_name in raw_room_map:
        room_map[sched_name] = raw_room_map[sched_name]
        matched += 1
    else:
        nk = normalize(sched_name)
        if nk in norm_room:
            room_map[sched_name] = norm_room[nk][1]
            matched += 1
print(f"Matched {matched}/{len(companies_list)} schedule companies to room numbers")

# Build JSON constants
raw_json  = json.dumps(students, ensure_ascii=False, separators=(',', ':'))
comp_json = json.dumps(companies_list, ensure_ascii=False, separators=(',', ':'))
room_json = json.dumps(room_map, ensure_ascii=False, separators=(',', ':'))
data_block = f"const RAW={raw_json};\nconst COMPANY={comp_json};\nconst ROOM_DATA_INIT={room_json};\nconst DEFAULT_COMP_PWD_INIT='MITM';\n"

# Read JS logic from separate files
with open('cloud_store_day2.js', 'r', encoding='utf-8') as f:
    cloud_store = f.read()

with open('portal_logic.js', 'r', encoding='utf-8') as f:
    js_logic = f.read()

with open('company_logic.js', 'r', encoding='utf-8') as f:
    company_logic = f.read()

full_script = data_block + "\n" + cloud_store + "\n" + js_logic + "\n" + company_logic

# Read Day 1 HTML as template
with open('index.html', 'r', encoding='utf-8') as f:
    html = f.read()

# Change title to Day 2
html = html.replace('<title>Placement Drive Portal 2026</title>', '<title>Placement Drive Portal 2026 — Day 2</title>')

# Update credentials table
old_start = html.find('<table class="creds-table">')
old_end   = html.find('</table>', old_start) + len('</table>')
if old_start != -1:
    rows = "\n".join(
        f"<tr><td class='us'>{c.lower().replace(' ','')[:20]}</td><td class='mu'>{c}</td></tr>"
        for c in companies_list
    )
    new_table = (
        "<table class=\"creds-table\">\n"
        "<tr><td><strong>Login ID</strong></td><td><strong>Company Name</strong></td></tr>\n"
        + rows + "\n</table>"
    )
    html = html[:old_start] + new_table + html[old_end:]

# Replace the script block
pattern = re.compile(r'<script>.*?</script>', re.DOTALL)
new_script_tag = f"<script>\n{full_script}\n</script>"

# Ensure Firebase SDK is present
firebase_sdk = '''    <script src="https://www.gstatic.com/firebasejs/10.12.0/firebase-app-compat.js"></script>
    <script src="https://www.gstatic.com/firebasejs/10.12.0/firebase-firestore-compat.js"></script>'''
if 'firebase-app-compat' not in html:
    html = html.replace('</head>', firebase_sdk + '\n</head>')

# Add loading overlay after <body> if not already present
loading_overlay = '''    <div id="loading-overlay" style="position:fixed;top:0;left:0;width:100%;height:100%;background:var(--bg,#090d1a);display:flex;align-items:center;justify-content:center;z-index:9999;flex-direction:column;gap:12px">
        <div style="width:36px;height:36px;border:3px solid rgba(99,102,241,.2);border-top:3px solid #6366f1;border-radius:50%;animation:spin 1s linear infinite"></div>
        <div style="font-family:'DM Sans',sans-serif;font-size:12px;color:#5c6494;letter-spacing:1px">LOADING PORTAL DATA...</div>
        <style>@keyframes spin{to{transform:rotate(360deg)}}</style>
    </div>'''
if 'id="loading-overlay"' not in html:
    html = html.replace('<body>', '<body>\n' + loading_overlay)

if pattern.search(html):
    html = pattern.sub(lambda _: new_script_tag, html)
else:
    print("Warning: <script> tag not found, attempting fallback...")
    ss = html.find('<script>')
    se = html.rfind('</script>')
    if ss != -1 and se != -1:
        html = html[:ss] + new_script_tag + html[se + len('</script>'):]

# ── DAY 2 SPECIFIC UI MODIFICATIONS ──────────────────────────────────────────

# 1. Replace Day 1 badge with Day 2 badge (if Day 1 badge exists from index.html)
html = html.replace(
    'linear-gradient(135deg,var(--ind),var(--pur));color:#fff;padding:4px 14px;border-radius:20px;font-size:11px;font-weight:700;font-family:var(--fnh);letter-spacing:.5px">&#128197; DAY 1</span>',
    'linear-gradient(135deg,var(--amb),#ef4444);color:#fff;padding:4px 14px;border-radius:20px;font-size:11px;font-weight:700;font-family:var(--fnh);letter-spacing:.5px">&#128197; DAY 2</span>'
)
# If no Day 1 badge, add Day 2 badge
if 'DAY 2' not in html:
    html = html.replace(
        'Centralized Interview Scheduling &amp; Tracking</div>',
        'Centralized Interview Scheduling &amp; Tracking</div>\n                <div style="margin-top:8px"><span style="background:linear-gradient(135deg,var(--amb),#ef4444);color:#fff;padding:4px 14px;border-radius:20px;font-size:11px;font-weight:700;font-family:var(--fnh);letter-spacing:.5px">&#128197; DAY 2</span></div>'
    )

# 2. Fix the Day switcher: make Day 2 active and Day 1 a link
#    If Day 1 switcher exists (from copied index.html), replace it
if 'class="day-btn active">&#128197; Day 1</span>' in html:
    html = html.replace(
        '<span class="day-btn active">&#128197; Day 1</span>',
        '<a href="index.html" class="day-btn">&#128197; Day 1</a>'
    )
    html = html.replace(
        '<a href="day2.html" class="day-btn">&#128197; Day 2</a>',
        '<span class="day-btn active">&#128197; Day 2</span>'
    )
elif 'class="day-switcher"' not in html:
    # No switcher at all — add one fresh
    day2_switcher = '''            <div class="day-switcher">
                <a href="index.html" class="day-btn">&#128197; Day 1</a>
                <span class="day-btn active">&#128197; Day 2</span>
            </div>
'''
    html = html.replace('<div class="rtabs">', day2_switcher + '            <div class="rtabs">')

# 3. Add Day switcher CSS if not already present
if '.day-switcher' not in html:
    day_switcher_css = '''
        .day-switcher {
            display: flex;
            gap: 8px;
            justify-content: center;
            margin-bottom: 18px;
        }
        .day-btn {
            padding: 8px 22px;
            border-radius: 20px;
            border: 1.5px solid var(--bd);
            background: var(--surf);
            color: var(--mut);
            font-family: var(--fnh);
            font-size: 11px;
            font-weight: 700;
            cursor: pointer;
            text-decoration: none;
            transition: all .2s;
            letter-spacing: .5px;
        }
        .day-btn:hover {
            border-color: var(--bd2);
            color: var(--tx);
        }
        .day-btn.active {
            background: linear-gradient(135deg, var(--ind), var(--pur));
            border-color: var(--ind);
            color: #fff;
            cursor: default;
        }
'''
    html = html.replace('</style>\n    <script src="https://cdnjs.cloudflare.com/ajax/libs/jspdf',
                         day_switcher_css + '\n    </style>\n    <script src="https://cdnjs.cloudflare.com/ajax/libs/jspdf')

# 4. Update the password hint for Day 2
html = html.replace(
    'Default password is your Date of Birth (DD/MM/YYYY)',
    'Default password for Day 2 is <b>01/01/2004</b>'
)

with open('day2.html', 'w', encoding='utf-8') as f:
    f.write(html)

print("Day 2 Portal rebuilt successfully!")
print(f"File size: {len(html.encode('utf-8')) / 1024:.0f} KB")

