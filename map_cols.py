
import openpyxl
wb = openpyxl.load_workbook('Master Copy of Students for Placement Drive.xlsx', data_only=True)
ws = wb.active
headers = [cell.value for cell in ws[1]]
data = [cell.value for cell in ws[2]]
with open('col_map.txt', 'w', encoding='utf-8') as f:
    for i, h in enumerate(headers):
        val = data[i] if i < len(data) else 'N/A'
        f.write(f"{i}: {h} -> {val}\n")
print("Mapping saved to col_map.txt")
